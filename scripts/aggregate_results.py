"""
aggregate_results.py — Tổng hợp toàn bộ kết quả JSON → mean±std → Excel/CSV
==============================================================================

Cách dùng:
    # Sau khi giải nén tar từ server vào results_raw/
    python scripts/aggregate_results.py --raw_dir results_raw/

    # Tùy chỉnh output
    python scripts/aggregate_results.py --raw_dir results_raw/ --out results_summary.xlsx

    # Chỉ in bảng ra terminal (không ghi Excel)
    python scripts/aggregate_results.py --raw_dir results_raw/ --no_excel

Workflow update kết quả mới:
    1. Chạy scripts/pack_results.sh trên server → tar mới
    2. SCP tar về, giải nén vào results_raw/ (overwrite OK)
    3. Chạy lại script này → Excel tự update

Cấu trúc đầu ra (Excel):
    Sheet "RetailRocket_full"       — tất cả methods, session full
    Sheet "RetailRocket_byLen"      — breakdown theo maxlen3/4/5/6
    Sheet "Diginetica_full"         — tương tự Diginetica
    Sheet "Diginetica_byLen"
    Sheet "All_raw"                 — toàn bộ seed-level data (dễ verify)
==============================================================================
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 numbers)
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("[WARN] openpyxl không có — chỉ xuất CSV. "
          "Cài: pip install openpyxl --break-system-packages")

# ── Method labels ─────────────────────────────────────────────────────────────
# Mapping: (ablation_field, folder_pattern) → (label, order)
# order dùng để sort trong bảng

METHOD_ORDER = {
    # Non-neural baselines
    "Popularity":   (0,  "Popularity"),
    "Content-KNN":  (1,  "Content-KNN"),
    # Warm-start neural baselines
    "A1":           (2,  "SR-GNN (A1)"),
    "GCEGNN":       (3,  "GCE-GNN"),
    "GCE-GNN":      (3,  "GCE-GNN"),
    "CORE":         (4,  "CORE"),
    "NCLSBR":       (5,  "NCL-SBR"),
    "NCL-SBR":      (5,  "NCL-SBR"),
    "CL4SRec":      (6,  "CL4SRec"),
    # Cold-start baselines
    "NirGNN":       (7,  "NirGNN"),
    "DC2R":         (8,  "DC2R"),
    "M2TRec":       (9,  "M2TRec"),
    # CatProCL ablations
    "A3":           (10, "A3 (+ EMA proto, no CL)"),
    "A4":           (11, "A4 (+ EMA proto + L_proto)"),
    "A7":           (15, "A7 CatProCL ★"),
    "A11":          (24, "CatPro-CL A11 ★"),
    "A11_p020":     (20, "A11 (λ=0.1, p=0.20)"),
    "A11_p015":     (21, "A11 (λ=0.1, p=0.15)"),
    "A11_p010":     (22, "A11 (λ=0.1, p=0.10)"),
    "A11_p008":     (23, "A11 (λ=0.1, p=0.08)"),
    "A11_p005":     (24, "A11 (λ=0.1, p=0.05)"),
    "A11_l030p005": (25, "A11 (λ=0.3, p=0.05) ★"),
    "A13":          (30, "A13 (Session-Proto NCE)"),
    "A14":          (31, "A14 (Target-EMA proto)"),
    "A15":          (32, "A15 (Additive Masked Rec)"),
    # v6 folder-name aliases (fallback khi JSON không có ablation field)
    "1_A11":        (24, "CatPro-CL A11 ★"),
    "2_M2TRec":     (9,  "M2TRec"),
    "3_DC2R":       (8,  "DC2R"),
    "4_NirGNN":     (7,  "NirGNN"),
}

METRICS_MAIN = ["HR@10", "HR@20", "MRR@10", "MRR@20"]
METRICS_COLD = ["Cold_HR@10", "Cold_HR@20", "Cold_MRR@10", "Cold_MRR@20"]
METRICS_ALL  = METRICS_MAIN + METRICS_COLD


# ── Infer method key from file path + JSON data ───────────────────────────────

def infer_method_key(fpath: Path, data: dict) -> str:
    """Return a canonical method key from file path + JSON config."""
    # v6 format: "model_key"; old: "ablation"; DC2R old: "model" (e.g. "DC2R_official")
    ablation = (data.get("model_key")
                or data.get("ablation")
                or data.get("model", "?").split("_")[0])  # "DC2R_official" → "DC2R"
    catprocl  = data.get("config", {}).get("catprocl", {})
    mask_prob  = catprocl.get("mask_prob",   None)
    lambda_v   = catprocl.get("lambda_proto", None)

    path_str = str(fpath).replace("\\", "/")

    # A11 variants — disambiguate by folder name (all have ablation="A11")
    if ablation == "A11":
        # Try to read actual params from JSON config first (most reliable)
        if mask_prob is not None and lambda_v is not None:
            lk = f"l{round(lambda_v*100):03d}"
            pk = f"p{round(mask_prob*100):03d}"
            return f"A11_{lk}{pk}"
        # Fallback: parse folder name
        m = re.search(r"results[_-](?:dig[_-])?a11[_-](l\d+[_-]p\d+|p\d+)",
                      path_str, re.IGNORECASE)
        if m:
            tag = m.group(1).replace("-", "_")
            return f"A11_{tag}"
        # v6 canonical A11 (no sweep variants) → use "A11" directly
        return "A11"

    # Standard ablations & baselines
    if ablation in METHOD_ORDER:
        return ablation

    # Heuristic for baselines with non-standard names
    abl_lower = ablation.lower()
    for known in ["m2trec", "gcegnn", "gce-gnn", "core", "nclsbr", "cl4srec",
                  "nirgnn", "popularity", "content-knn"]:
        if known.replace("-", "") in abl_lower.replace("-", ""):
            return ablation  # keep as-is, will be sorted last

    return ablation


def get_method_label(method_key: str) -> tuple[int, str]:
    """Return (sort_order, display_label) for a method key."""
    if method_key in METHOD_ORDER:
        return METHOD_ORDER[method_key]

    # A11 variants: parse from key like "A11_l030p005"
    m = re.match(r"A11_(l(\d+)p(\d+)|p(\d+))", method_key)
    if m:
        if m.group(2):   # has lambda
            lam  = int(m.group(2)) / 100
            prob = int(m.group(3)) / 100
            label = f"A11 (λ={lam:.2g}, p={prob:.2g})"
            # sort order based on lambda + mask
            order = 20 + int(m.group(2))/10 + int(m.group(3))/1000
        else:            # lambda=0.1 implied (old sweep)
            prob = int(m.group(4)) / 100
            label = f"A11 (λ=0.1, p={prob:.2g})"
            order = 20 + 0.1 + int(m.group(4))/1000
        return (int(order*10), label)

    return (99, method_key)


def get_dataset_base(dataset: str) -> str:
    for d in ("retailrocket", "diginetica", "yoochoose", "cellphones"):
        if d in dataset.lower():
            return d
    return dataset.split("_")[0]


def get_session_len(dataset: str) -> str:
    m = re.search(r"maxlen(\d+)", dataset)
    return f"maxlen{m.group(1)}" if m else "full"


# ── Metrics extraction (handles both v6 and old DC2R format) ─────────────────

def extract_metrics(data: dict) -> dict:
    """Handle two JSON formats:
    - v6 / standard: data["test"]["HR@20"], data["test"]["Cold_HR@20"]
    - Old DC2R:       data["results"]["overall"]["HR@20"], data["results"]["cold"]["HR@20"]
    Returns flat dict with all METRICS_ALL keys (None if missing).
    """
    if "test" in data:
        t = data["test"]
        return {m: t.get(m) for m in METRICS_ALL}
    elif "results" in data:
        ov   = data["results"].get("overall", {})
        cold = data["results"].get("cold", {})
        return {
            "HR@10":       ov.get("HR@10"),
            "HR@20":       ov.get("HR@20"),
            "MRR@10":      ov.get("MRR@10"),
            "MRR@20":      ov.get("MRR@20"),
            "Cold_HR@10":  cold.get("HR@10"),
            "Cold_HR@20":  cold.get("HR@20"),
            "Cold_MRR@10": cold.get("MRR@10"),
            "Cold_MRR@20": cold.get("MRR@20"),
        }
    return {m: None for m in METRICS_ALL}


# ── Load all JSON files ───────────────────────────────────────────────────────

def load_all_jsons(raw_dir: Path) -> list[dict]:
    """Recursively load all JSON result files from raw_dir."""
    records = []
    skipped = 0
    for fpath in sorted(raw_dir.rglob("*.json")):
        try:
            with open(fpath, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"  [WARN] Skip {fpath.name}: {e}")
            skipped += 1
            continue

        # Extract metrics (handles both formats)
        metrics = extract_metrics(data)
        if metrics.get("HR@20") is None and metrics.get("HR@10") is None:
            skipped += 1
            continue

        dataset    = data.get("dataset", fpath.stem.split("_")[0])
        # v6: "model_key"; old: "ablation"; DC2R old: "model" (e.g. "DC2R_official")
        ablation   = (data.get("model_key")
                      or data.get("ablation")
                      or data.get("model", "?").split("_")[0])
        seed       = data.get("seed", -1)
        best_epoch = data.get("best_epoch", -1)
        # v6 has explicit session_len field; old DC2R has neither — infer from path
        session_len = data.get("session_len") or get_session_len(dataset)
        if not session_len or session_len in ("full", "fullen", ""):
            # Try to infer from folder path (e.g. .../maxlen3/3_DC2R/...)
            m_path = re.search(r'[\\/](maxlen\d+|fullen)[\\/]', str(fpath))
            if m_path:
                session_len = m_path.group(1)
        # Normalize "full" → "fullen" for v6 consistency
        if session_len == "full":
            session_len = "fullen"

        method_key   = infer_method_key(fpath, data)
        dataset_base = get_dataset_base(dataset)

        rec = {
            "fpath":        str(fpath),
            "dataset":      dataset,
            "dataset_base": dataset_base,
            "session_len":  session_len,
            "method_key":   method_key,
            "ablation":     ablation,
            "seed":         seed,
            "best_epoch":   best_epoch,
        }
        for m in METRICS_ALL:
            rec[m] = metrics.get(m)

        # Harmonic mean Overall HR@20 & Cold HR@20
        ov = metrics.get("HR@20")
        co = metrics.get("Cold_HR@20")
        if ov and co and (ov + co) > 0:
            rec["H_HR@20"] = 2 * ov * co / (ov + co)
        else:
            rec["H_HR@20"] = None

        records.append(rec)

    print(f"  Loaded {len(records)} records, skipped {skipped}")
    return records


# ── Aggregate: mean ± std ─────────────────────────────────────────────────────

def aggregate(records: list[dict]) -> list[dict]:
    """Group by (dataset_base, session_len, method_key), compute mean±std."""
    import statistics

    groups: dict = defaultdict(list)
    for r in records:
        key = (r["dataset_base"], r["session_len"], r["method_key"])
        groups[key].append(r)

    agg_rows = []
    for (db, sl, mk), recs in groups.items():
        seeds = sorted(set(r["seed"] for r in recs))
        order, label = get_method_label(mk)
        row = {
            "dataset_base": db,
            "session_len":  sl,
            "method_key":   mk,
            "method_label": label,
            "sort_order":   order,
            "n_seeds":      len(recs),
            "seeds":        str(seeds),
        }
        for m in METRICS_ALL + ["H_HR@20"]:
            vals = [r[m] for r in recs if r.get(m) is not None]
            if vals:
                mean = statistics.mean(vals)
                std  = statistics.stdev(vals) if len(vals) > 1 else 0.0
                row[f"{m}_mean"] = mean
                row[f"{m}_std"]  = std
                row[f"{m}_fmt"]  = f"{mean*100:.2f} ± {std*100:.2f}"
            else:
                row[f"{m}_mean"] = None
                row[f"{m}_std"]  = None
                row[f"{m}_fmt"]  = "—"
        agg_rows.append(row)

    # Sort
    agg_rows.sort(key=lambda r: (
        r["dataset_base"],
        r["session_len"],
        r["sort_order"]
    ))
    return agg_rows


# ── Print terminal table ──────────────────────────────────────────────────────

def print_table(agg: list[dict], dataset_base: str = None, session_len: str = None):
    rows = agg
    if dataset_base:
        rows = [r for r in rows if r["dataset_base"] == dataset_base]
    if session_len:
        rows = [r for r in rows if r["session_len"] == session_len]
    if not rows:
        return

    print(f"\n{'='*115}")
    print(f"  {dataset_base or 'ALL'} | {session_len or 'ALL LENGTHS'}")
    print(f"{'='*115}")
    hdr = (f"{'Method':<35} {'n':>2} | "
           f"{'HR@20':>7} {'ColdHR@20':>10} {'H':>7} | "
           f"{'HR@10':>7} {'MRR@20':>7} {'ColdMRR@20':>11}")
    print(hdr)
    print("-" * 115)

    cur_sl = None
    for r in rows:
        if r["session_len"] != cur_sl:
            cur_sl = r["session_len"]
            print(f"\n  ── {cur_sl} ──")
        n = r["n_seeds"]
        print(
            f"  {r['method_label']:<33} {n:>2}  | "
            f"{r.get('HR@20_fmt','—'):>17} {r.get('Cold_HR@20_fmt','—'):>17} "
            f"{r.get('H_HR@20_fmt','—'):>14} | "
            f"{r.get('HR@10_fmt','—'):>17} {r.get('MRR@20_fmt','—'):>17} "
            f"{r.get('Cold_MRR@20_fmt','—'):>17}"
        )
    print("=" * 115)


# ── Excel output ──────────────────────────────────────────────────────────────

# Styling constants
C_HEADER  = "1F4E79"  # dark blue header
C_SUBHDR  = "2E75B6"  # medium blue subheader
C_BEST    = "E2EFDA"  # light green for best method rows
C_ODD     = "F5F5F5"  # light gray alternating rows
C_WARN    = "FFF2CC"  # yellow for incomplete (< 5 seeds)

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def xl_header(ws, row, col, text, bold=True, bg=C_HEADER, color="FFFFFF",
              wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=color, name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=wrap)
    c.border = BORDER
    return c


def xl_cell(ws, row, col, value, bold=False, bg=None, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c


def write_summary_sheet(wb, sheet_name: str, agg: list[dict],
                        dataset_base: str, session_len_filter: str = None):
    """Write one summary sheet: rows=methods, cols=metrics (mean±std as strings)."""
    rows = [r for r in agg if r["dataset_base"] == dataset_base]
    if session_len_filter:
        rows = [r for r in rows if r["session_len"] == session_len_filter]
    if not rows:
        return

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    DISPLAY_COLS = [
        ("HR@20",        "Overall\nHR@20"),
        ("HR@10",        "Overall\nHR@10"),
        ("MRR@20",       "Overall\nMRR@20"),
        ("MRR@10",       "Overall\nMRR@10"),
        ("Cold_HR@20",   "Cold\nHR@20"),
        ("Cold_HR@10",   "Cold\nHR@10"),
        ("Cold_MRR@20",  "Cold\nMRR@20"),
        ("Cold_MRR@10",  "Cold\nMRR@10"),
        ("H_HR@20",      "H\n(Harmonic)"),
    ]

    # Row 1: sheet title
    title = f"{dataset_base.title()} — {session_len_filter or 'All Lengths'} | mean ± std (×100%)"
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=3 + len(DISPLAY_COLS))
    xl_header(ws, 1, 1, title, bg=C_HEADER, color="FFFFFF")
    ws.row_dimensions[1].height = 22

    # Row 2: column headers
    xl_header(ws, 2, 1, "Method",      bg=C_SUBHDR, color="FFFFFF")
    xl_header(ws, 2, 2, "n\nSeeds",   bg=C_SUBHDR, color="FFFFFF", wrap=True)
    xl_header(ws, 2, 3, "Session\nLen", bg=C_SUBHDR, color="FFFFFF", wrap=True)
    for ci, (mk, hdr) in enumerate(DISPLAY_COLS, start=4):
        xl_header(ws, 2, ci, hdr, bg=C_SUBHDR, color="FFFFFF", wrap=True)
    ws.row_dimensions[2].height = 34

    # Data rows
    cur_sl = None
    data_row = 3
    for r in rows:
        is_section_start = (r["session_len"] != cur_sl)
        if is_section_start and session_len_filter is None:
            cur_sl = r["session_len"]
            # Section separator row
            ws.merge_cells(start_row=data_row, start_column=1,
                           end_row=data_row, end_column=3 + len(DISPLAY_COLS))
            c = ws.cell(row=data_row, column=1,
                        value=f"── {cur_sl} ──")
            c.font = Font(bold=True, italic=True, color="1F4E79",
                          name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor="D6E4F0")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        # Highlight best rows (A7 and A11 l030_p005)
        is_starred = ("★" in r["method_label"])
        bg = C_BEST if is_starred else (C_ODD if data_row % 2 == 0 else None)
        warn = r["n_seeds"] < 5

        xl_cell(ws, data_row, 1, r["method_label"], bold=is_starred,
                bg=C_WARN if warn else bg, align="left")
        xl_cell(ws, data_row, 2, r["n_seeds"],
                bg=C_WARN if warn else bg)
        xl_cell(ws, data_row, 3, r["session_len"],
                bg=bg)

        for ci, (mk, _) in enumerate(DISPLAY_COLS, start=4):
            val = r.get(f"{mk}_fmt", "—")
            xl_cell(ws, data_row, ci, val, bold=is_starred,
                    bg=C_BEST if is_starred else bg)

        data_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 9
    for ci in range(4, 4 + len(DISPLAY_COLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    ws.freeze_panes = "A3"


def write_raw_sheet(wb, records: list[dict]):
    """Write all seed-level raw data for verification."""
    ws = wb.create_sheet("All_raw")
    ws.sheet_view.showGridLines = False

    COLS = ["dataset_base", "session_len", "method_label", "seed", "best_epoch",
            "HR@20", "HR@10", "MRR@20", "MRR@10",
            "Cold_HR@20", "Cold_HR@10", "Cold_MRR@20", "Cold_MRR@10",
            "H_HR@20"]
    HEADERS = {
        "dataset_base": "Dataset",
        "session_len": "SessLen",
        "method_label": "Method",
        "seed": "Seed",
        "best_epoch": "Epoch",
        "HR@20": "HR@20",
        "HR@10": "HR@10",
        "MRR@20": "MRR@20",
        "MRR@10": "MRR@10",
        "Cold_HR@20": "Cold HR@20",
        "Cold_HR@10": "Cold HR@10",
        "Cold_MRR@20": "Cold MRR@20",
        "Cold_MRR@10": "Cold MRR@10",
        "H_HR@20": "H (Harmonic)",
    }

    for ci, col in enumerate(COLS, start=1):
        xl_header(ws, 1, ci, HEADERS[col], bg=C_SUBHDR, color="FFFFFF")

    # Sort records for readability
    sorted_recs = sorted(
        records,
        key=lambda r: (r["dataset_base"], r["session_len"],
                       get_method_label(r["method_key"])[0], r["seed"])
    )

    for ri, r in enumerate(sorted_recs, start=2):
        bg = C_ODD if ri % 2 == 0 else None
        for ci, col in enumerate(COLS, start=1):
            val = r.get(col)
            if val is None:
                val = "—"
            elif col in ("HR@20", "HR@10", "MRR@20", "MRR@10",
                         "Cold_HR@20", "Cold_HR@10", "Cold_MRR@20", "Cold_MRR@10",
                         "H_HR@20"):
                if isinstance(val, float):
                    val = round(val * 100, 4)
            xl_cell(ws, ri, ci, val, bg=bg)

    col_widths = {"A": 14, "B": 9, "C": 36, "D": 6, "E": 6}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for ci in range(6, 6 + 9):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"


def write_excel(agg: list[dict], records: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    datasets = sorted(set(r["dataset_base"] for r in agg))
    for db in datasets:
        write_summary_sheet(wb, f"{db.title()}_full",
                            agg, db, session_len_filter="fullen")
        write_summary_sheet(wb, f"{db.title()}_byLen",
                            agg, db, session_len_filter=None)

    write_raw_sheet(wb, records)
    wb.save(str(out_path))
    print(f"  Excel saved → {out_path}")


# ── CSV output ────────────────────────────────────────────────────────────────

def write_csv(agg: list[dict], out_path: Path):
    import csv
    cols = ["dataset_base", "session_len", "method_label", "n_seeds"]
    for m in METRICS_ALL + ["H_HR@20"]:
        cols += [f"{m}_mean", f"{m}_std", f"{m}_fmt"]

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for row in agg:
            writer.writerow({k: row.get(k, "") for k in cols})
    print(f"  CSV saved  → {out_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Aggregate CatPro-CL result JSONs → mean±std → Excel")
    parser.add_argument("--raw_dir",   default="results_raw",
                        help="Thư mục chứa JSON đã giải nén từ server (default: results_raw/)")
    parser.add_argument("--out",       default=None,
                        help="Output Excel file (default: results_raw/../results_summary.xlsx)")
    parser.add_argument("--no_excel",  action="store_true",
                        help="Chỉ in terminal, không ghi Excel")
    parser.add_argument("--dataset",   default=None,
                        help="Filter dataset để in (vd: retailrocket)")
    parser.add_argument("--len",       default="full",
                        help="Session length để in terminal (default: full)")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    if not raw_dir.exists():
        print(f"ERROR: raw_dir không tồn tại: {raw_dir}")
        sys.exit(1)

    print(f"\nLoading JSONs from: {raw_dir}")
    records = load_all_jsons(raw_dir)
    if not records:
        print("Không tìm thấy JSON nào!")
        sys.exit(1)

    print(f"\nAggregating {len(records)} records...")
    agg = aggregate(records)

    # Terminal output
    datasets = sorted(set(r["dataset_base"] for r in agg))
    for db in (([args.dataset] if args.dataset else datasets)):
        print_table(agg, dataset_base=db, session_len=args.len)

    if args.no_excel:
        return

    if not HAS_EXCEL:
        print("\n[INFO] pip install openpyxl để xuất Excel")
        out_csv = Path(args.out).with_suffix(".csv") if args.out else raw_dir.parent / "results_summary.csv"
        write_csv(agg, out_csv)
        return

    out_path = Path(args.out) if args.out else raw_dir.parent / "results_summary.xlsx"
    write_excel(agg, records, out_path)

    # Also write CSV (git-trackable, no binary)
    write_csv(agg, out_path.with_suffix(".csv"))

    print(f"\nDone. Mở {out_path} để xem kết quả tổng hợp.")


if __name__ == "__main__":
    main()
